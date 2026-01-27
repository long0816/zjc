import streamlit as st
import pandas as pd
import io
import re
import time
import random
import xlsxwriter
import traceback
import openpyxl
from datetime import datetime, timedelta
import altair as alt

# 尝试导入解密库
try:
    import msoffice_crypto

    HAS_CRYPT = True
except ImportError:
    HAS_CRYPT = False


# ================= 0. 样式配置 =================
def inject_custom_css():
    st.markdown("""
        <style>
            .stApp { background-color: #f4f6f9; }
            [data-testid="stMetricValue"] { font-size: 26px; color: #2E86C1; font-weight: bold; }
            div[data-testid="stDataFrame"] { font-size: 14px; border: 1px solid #ddd; }
            div.stButton > button {
                width: 100%; border-radius: 6px; height: 42px; font-weight: 600; border: none;
                background-color: #34495E; color: white; transition: 0.2s;
            }
            div.stButton > button:hover { background-color: #2C3E50; color: #F39C12; transform: translateY(-2px); }
            .info-box {
                padding: 15px; background-color: #E8F6F3; border-left: 5px solid #1ABC9C;
                border-radius: 4px; color: #2C3E50; margin-bottom: 15px;
            }
        </style>
    """, unsafe_allow_html=True)


# ================= 1. 基础工具函数 =================

def decrypt_file(file_obj, password):
    if not HAS_CRYPT: raise ImportError("请先安装解密模块: pip install msoffice-crypto-tool")
    file_obj.seek(0)
    decrypted = io.BytesIO()
    office_file = msoffice_crypto.OfficeFile(file_obj)
    office_file.load_key(password=password)
    office_file.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted


def clean_date(date_val):
    """ 日期清洗函数 """
    if pd.isna(date_val): return pd.NaT
    val_str = str(date_val).strip()
    if val_str == "" or val_str == "0" or val_str.lower() in ['nan', 'none', '-', '/']:
        return pd.NaT
    if isinstance(date_val, (datetime, pd.Timestamp)):
        if date_val.year < 2020: return pd.NaT
        return pd.to_datetime(date_val)
    if re.match(r'^\d+(\.\d+)?$', val_str):
        try:
            val_float = float(val_str)
            if val_float < 43831: return pd.NaT
            return pd.to_datetime(val_float, unit='D', origin='1899-12-30')
        except:
            pass
    clean_str = val_str.replace('.', '-').replace('/', '-').replace('\\', '-')
    try:
        dt = pd.to_datetime(clean_str, errors='coerce')
        if pd.notna(dt) and dt.year < 2020: return pd.NaT
        return dt
    except:
        return pd.NaT


def extract_diffusion_lot(wafer_id):
    if pd.isna(wafer_id): return ""
    wafer_id = str(wafer_id).upper().strip()
    parts = re.split(r'[-.]', wafer_id)
    return parts[0] if parts else wafer_id


def determine_cassette_info(cid):
    if pd.isna(cid) or str(cid).strip() in ['-', 'nan', '', 'None']: return "未知/空", "无"
    cid = str(cid).upper().strip()
    prefix_match = re.match(r'([A-Z-]+)', cid)
    prefix = prefix_match.group(1) if prefix_match else "未知"
    if cid.endswith("-C") or any(x in cid for x in ["TS-C", "TZ-C", "WAP-C", "HPC-C"]): return "Cassette(片盒)", prefix
    if cid.endswith("-B") or any(x in cid for x in ["HPC-B", "WAP-B"]): return "八角盒", prefix
    if cid.startswith("TS") or cid.startswith("FO"): return "13槽 Foup", prefix
    if any(cid.startswith(x) for x in ["HPC", "WAP", "BU"]): return "25槽 Foup", prefix
    return "FO料盒", prefix


def find_header_row(df, search_rows=20):
    key_map = {
        'PRODUCT': ['PRODUCT ID', 'PRODUCT', 'TP NO', 'TP NUMBER', '工单', 'TP #'],
        'LOT': ['LOT ID', 'LOT NO', 'BATCH ID', '作业工单', 'LOT #'],
        'WAFER': ['WAFER ID', 'CHIP ID', '芯片号', '芯片'],
        'CASSETTE': ['CASSETTE', '料盒'],
        'DATE': ['TIME', 'DATE', '日期', '时间']
    }
    for i in range(min(len(df), search_rows)):
        row_values = [str(x).upper().strip() for x in df.iloc[i].tolist()]
        row_str = " ".join(row_values)
        score = 0
        if any(k in row_str for k in key_map['PRODUCT']): score += 1
        if any(k in row_str for k in key_map['LOT']): score += 1
        if any(k in row_str for k in key_map['WAFER']): score += 1
        if score >= 2: return i
    return 0


def identify_system_columns(df):
    sys_cols = {}
    target_keywords = {
        '_sys_product': ['PRODUCT', 'TP NO', '工单'],
        '_sys_lot': ['LOT ID', '作业工单', '批号'],
        '_sys_wafer': ['WAFER', 'CHIP ID', '芯片'],
        '_sys_cassette': ['CASSETTE', '料盒', 'BOX ID'],
        '_sys_location': ['LOCATION', '库位'],
        '_sys_in_date': ['接收', '入库', 'IN_TIME'],
        '_sys_out_date': ['领用', '出库', 'OUT_TIME'],
        '_sys_type_raw': ['选择', 'SELECT', 'TYPE'],
        '_sys_remark': ['备注', 'REMARK']
    }
    used_cols = set()
    for sys_key, keywords in target_keywords.items():
        for col in df.columns:
            if col in used_cols: continue
            c_upper = str(col).strip().upper()
            if any(k in c_upper for k in keywords):
                sys_cols[col] = sys_key
                used_cols.add(col)
                break
    return sys_cols


def determine_chip_type(row, file_name, sheet_name):
    if '_sys_type_raw' in row and pd.notna(row['_sys_type_raw']):
        val = str(row['_sys_type_raw']).strip().upper()
        if "REAL" in val: return "Real"
        if "不可" in val or "NG" in val or "NON" in val: return "不可回货"
        if "DUMMY" in val: return "Dummy"
    f_upper = file_name.upper();
    s_upper = sheet_name.upper()
    if "DUMMY" in f_upper: return "Dummy"
    if "REAL" in s_upper: return "Real"
    if "不可" in s_upper or "NG" in s_upper: return "不可回货"
    if "DUMMY" in s_upper: return "Dummy"
    return "Real"


def get_client_name(file_name, sheet_name):
    f_name_no_ext = file_name.rsplit('.', 1)[0]
    f_upper = f_name_no_ext.upper()
    s_upper = sheet_name.strip().upper()
    if "E20" in f_upper: return "E20"
    if "DUMMY" in f_upper or "231" in f_upper: return s_upper.replace("BX", "").strip()
    if "中间仓" in f_name_no_ext:
        client_part = f_name_no_ext.split("中间仓")[0]
        return client_part.strip().upper()
    return f_upper


def check_grey_shipped(file_obj, sheet_name, header_idx, df_len, use_grey_logic):
    """
    灰色行识别逻辑
    如果 use_grey_logic=False，则直接返回全False，不浪费性能
    """
    if not use_grey_logic: return [False] * df_len

    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb[sheet_name]
        is_grey_list = []
        start_row = header_idx + 2

        curr_idx = 0
        for row in ws.iter_rows(min_row=start_row, max_row=start_row + df_len - 1):
            if curr_idx >= df_len: break
            cell_a = row[0]
            is_grey = False
            if cell_a.fill and cell_a.fill.patternType == 'solid':
                color = cell_a.fill.start_color
                if color.index != '00000000' and color.index != 'FFFFFFFF':
                    if not (hasattr(color, 'rgb') and color.rgb == '00000000'):
                        is_grey = True
            is_grey_list.append(is_grey)
            curr_idx += 1
        if len(is_grey_list) < df_len: is_grey_list.extend([False] * (df_len - len(is_grey_list)))
        return is_grey_list
    except:
        return [False] * df_len


@st.cache_data(show_spinner=False)
def process_uploaded_files(uploaded_files, password, use_grey_logic):
    all_data = []
    sheet_headers = {}
    loader = st.empty()
    funny_loading = [
        {"msg": "正在执行【只看字，不看色】法则... ⚖️",
         "gif": "https://media.giphy.com/media/l0HlHFRb68qGNz670/giphy.gif"},
        {"msg": "正在过滤【蓝膜】... 🔵", "gif": "https://media.giphy.com/media/13FrpeVH09Zrb2/giphy.gif"},
        {"msg": "备注有日期？不！那是在库！📅", "gif": "https://media.giphy.com/media/26n6WywJyh39n1pW8/giphy.gif"},
    ]

    for idx, file in enumerate(uploaded_files):
        pick = random.choice(funny_loading)
        loader.markdown(f"""
            <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 60px;">
                <img src="{pick['gif']}" width="500" style="border-radius: 15px; margin-bottom: 25px; box-shadow: 0 4px 20px rgba(0,0,0,0.3);">
                <h2 style="color: #E74C3C; font-family: 'Comic Sans MS', sans-serif; font-size: 28px; text-shadow: 1px 1px 2px #ddd;">{pick['msg']}</h2>
                <p style="color: #7F8C8D; font-size: 16px;">正在处理: {file.name}</p>
                <div style="width: 600px; background: #ddd; height: 20px; border-radius: 10px; overflow: hidden; margin-top: 15px;">
                    <div style="width: {int((idx / len(uploaded_files)) * 100)}%; background: linear-gradient(90deg, #E74C3C, #F1C40F); height: 100%; transition: width 0.3s;"></div>
                </div>
            </div>""", unsafe_allow_html=True)
        file.seek(0)
        file_bytes = file.read()
        file_obj = io.BytesIO(file_bytes)
        try:
            excel_file = pd.ExcelFile(file_obj)
        except:
            if password:
                try:
                    file_obj.seek(0)
                    decrypted = decrypt_file(file_obj, password)
                    excel_file = pd.ExcelFile(decrypted)
                    file_obj = decrypted
                except:
                    continue
            else:
                continue

        for sheet_name in excel_file.sheet_names:
            try:
                s_upper = str(sheet_name).strip().upper()
                f_upper = file.name.upper()
                is_valid = False
                if "在库库存" in s_upper or "出库库存" in s_upper:
                    is_valid = True
                elif "REAL" in s_upper or "不可回货" in s_upper or "DUMMY" in f_upper:
                    if "汇总" not in s_upper and "功能" not in s_upper: is_valid = True
                if "中间仓" in f_upper and ("汇总" not in s_upper and "功能" not in s_upper): is_valid = True
                if not is_valid: continue

                df_temp = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=20)
                header_idx = find_header_row(df_temp)
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_idx)

                if "E20" in f_upper and header_idx > 0 and len(df.columns) > 0:
                    c0 = df.columns[0]
                    if "Unnamed" in str(c0) or "选择" in str(c0): df.rename(columns={c0: "选择"}, inplace=True)

                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                df = df.loc[:, ~df.columns.duplicated()]

                clean_fname = file.name.rsplit('.', 1)[0]
                source_key = f"{clean_fname}-{sheet_name}"[:40]
                original_columns = df.columns.tolist()
                sheet_headers[source_key] = original_columns

                sys_map = identify_system_columns(df)
                if not any(k in sys_map.values() for k in ['_sys_product', '_sys_lot']): continue
                for original_col, sys_key in sys_map.items(): df[sys_key] = df[original_col]

                needed_sys = ['_sys_product', '_sys_lot', '_sys_wafer', '_sys_cassette', '_sys_location',
                              '_sys_in_date', '_sys_out_date', '_sys_type_raw', '_sys_remark']
                for ns in needed_sys:
                    if ns not in df.columns: df[ns] = pd.NaT if 'date' in ns else '-'

                if '_sys_remark' in df.columns:
                    # 🟢 如果 use_grey_logic 为 False，这里直接返回全False，不消耗时间
                    is_grey_shipped = check_grey_shipped(file_obj, sheet_name, header_idx, len(df),
                                                         use_grey_logic=use_grey_logic)
                    df['_sys_is_grey'] = is_grey_shipped
                else:
                    df['_sys_is_grey'] = False

                # 🟢🟢🟢 V7.0 核心修正 🟢🟢🟢
                # 1. 保留原始数据
                df['_sys_out_date_raw'] = df['_sys_out_date']

                # 2. 清洗日期 (用于统计)
                df['_sys_out_date'] = df['_sys_out_date'].apply(clean_date)

                # 3. 判定 A: 领用列有内容 -> 出库 (最高优先级)
                mask_has_content = df['_sys_out_date_raw'].apply(
                    lambda x: pd.notna(x) and str(x).strip() not in ['', '0', 'nan', 'None', '-'])

                # 4. 判定 B: 备注里有"蓝膜" -> 出库
                keywords = ['蓝膜', 'blue tape']
                df['_sys_remark'] = df['_sys_remark'].astype(str)
                mask_keyword = df['_sys_remark'].str.contains('|'.join(keywords), na=False, case=False)

                # 5. 判定 C: 灰色背景 (仅当侧边栏开启时生效，默认不看)
                mask_grey = df['_sys_is_grey']

                # 6. 综合判定 (任一满足即出库)
                mask_force_out = (
                        (mask_has_content & pd.isna(df['_sys_out_date'])) |  # 领用列有字 (即使解析失败)
                        (mask_keyword & pd.isna(df['_sys_out_date'])) |  # 备注有蓝膜
                        (mask_grey & pd.isna(df['_sys_out_date']))  # 灰色 (仅在开启时为True)
                )

                # 7. 赋予当前时间作为标记
                df.loc[mask_force_out, '_sys_out_date'] = pd.Timestamp.now()
                # 🟢🟢🟢 修改结束 🟢🟢🟢

                df['_sys_diff_lot'] = df['_sys_wafer'].apply(extract_diffusion_lot)
                df['_sys_in_date'] = df['_sys_in_date'].apply(clean_date)
                df['_sys_client'] = get_client_name(file.name, sheet_name)
                df['_sys_chip_type'] = df.apply(lambda row: determine_chip_type(row, file.name, sheet_name), axis=1)
                df['_sys_cass_type'], df['_sys_cass_prefix'] = zip(*df['_sys_cassette'].apply(determine_cassette_info))

                df['_sys_source_key'] = source_key
                all_data.append(df)
            except:
                pass

    loader.empty()
    if all_data: return pd.concat(all_data, ignore_index=True), sheet_headers
    return pd.DataFrame(), {}


# ================= 3. 统计逻辑 =================

def calculate_statistics(df, start_date, end_date):
    t_start = pd.to_datetime(start_date)
    t_end = pd.to_datetime(end_date) + timedelta(days=0.99999)
    l_end = t_start - timedelta(seconds=1)
    l_start = l_end - timedelta(days=(t_end - t_start).days)

    stats_in, stats_out, stats_stock = [], [], []
    for client, group in df.groupby('_sys_client'):
        in_this = group[(group['_sys_in_date'] >= t_start) & (group['_sys_in_date'] <= t_end)]
        in_last = group[(group['_sys_in_date'] >= l_start) & (group['_sys_in_date'] <= l_end)]
        out_this = group[(group['_sys_out_date'] >= t_start) & (group['_sys_out_date'] <= t_end)]
        out_last = group[(group['_sys_out_date'] >= l_start) & (group['_sys_out_date'] <= l_end)]
        stock_curr = group[pd.isna(group['_sys_out_date'])]

        def calc_growth(curr, last):
            if last == 0: return 0 if curr == 0 else 1.0
            return (curr - last) / last

        stats_in.append({'客户': client, '本期入库(片)': len(in_this), '上期入库(片)': len(in_last),
                         '差异': len(in_this) - len(in_last), '环比增长': calc_growth(len(in_this), len(in_last)),
                         '本期入库(批)': in_this['_sys_lot'].nunique()})
        stats_out.append({'客户': client, '本期出库(片)': len(out_this), '上期出库(片)': len(out_last),
                          '差异': len(out_this) - len(out_last), '环比增长': calc_growth(len(out_this), len(out_last)),
                          '本期出库(批)': out_this['_sys_lot'].nunique()})
        stats_stock.append(
            {'客户': client, '当前在库(片)': len(stock_curr), '当前在库(批)': stock_curr['_sys_lot'].nunique()})

    all_dates = pd.concat([df['_sys_in_date'], df['_sys_out_date']]).dropna()
    if all_dates.empty:
        return pd.DataFrame(stats_in), pd.DataFrame(stats_out), pd.DataFrame(stats_stock), pd.DataFrame()

    min_date, max_date = all_dates.min(), all_dates.max()
    full_range = pd.date_range(start=min_date.replace(day=1), end=max_date.replace(day=1) + pd.offsets.MonthBegin(1),
                               freq='MS')

    valid_in = df[pd.notna(df['_sys_in_date'])].copy()
    valid_in['Month'] = valid_in['_sys_in_date'].dt.to_period('M').dt.to_timestamp()
    monthly_in = valid_in.groupby('Month').size().reindex(full_range, fill_value=0)

    valid_out = df[pd.notna(df['_sys_out_date'])].copy()
    valid_out['Month'] = valid_out['_sys_out_date'].dt.to_period('M').dt.to_timestamp()
    monthly_out = valid_out.groupby('Month').size().reindex(full_range, fill_value=0)

    df_monthly_total = pd.DataFrame({
        '月份': full_range,
        '入库总数': monthly_in.values,
        '出库总数': monthly_out.values
    })
    df_monthly_total['月份_Str'] = df_monthly_total['月份'].dt.strftime('%Y-%m')

    return (pd.DataFrame(stats_in).fillna(0), pd.DataFrame(stats_out).fillna(0), pd.DataFrame(stats_stock).fillna(0),
            df_monthly_total)


# ================= 4. 高级报表生成 =================

def create_statistical_excel(df_in, df_out, df_stock, df_month, t_s, t_e):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True, 'nan_inf_to_errors': True})
    font = '微软雅黑'

    fmt_h_blue = workbook.add_format(
        {'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1, 'align': 'center',
         'valign': 'vcenter', 'font_name': font})
    fmt_h_org = workbook.add_format(
        {'bold': True, 'bg_color': '#ED7D31', 'font_color': 'white', 'border': 1, 'align': 'center',
         'valign': 'vcenter', 'font_name': font})
    fmt_h_grn = workbook.add_format(
        {'bold': True, 'bg_color': '#70AD47', 'font_color': 'white', 'border': 1, 'align': 'center',
         'valign': 'vcenter', 'font_name': font})
    fmt_data = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_name': font})
    fmt_pct = workbook.add_format(
        {'border': 1, 'align': 'center', 'valign': 'vcenter', 'num_format': '0.00%', 'font_name': font})
    fmt_red = workbook.add_format(
        {'font_color': '#D00000', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'num_format': '0.00%',
         'bold': True, 'font_name': font})
    fmt_green = workbook.add_format(
        {'font_color': '#008000', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'num_format': '0.00%',
         'bold': True, 'font_name': font})
    curr_str = f"{t_s.strftime('%m-%d')}~{t_e.strftime('%m-%d')}"

    def write_sheet(ws, title, df, header_fmt, chart_color):
        ws.merge_range('A1:F1', f'{title} ({curr_str})', header_fmt)
        cols = list(df.columns)
        ws.write_row('A2', cols, header_fmt)
        if df.empty: return
        for i, r in enumerate(df.to_dict('records')):
            row = i + 2
            ws.write(row, 0, r['客户'], fmt_data)
            ws.write(row, 1, r[cols[1]], fmt_data)
            ws.write(row, 2, r[cols[2]], fmt_data)
            ws.write(row, 3, r['差异'], fmt_data)
            val = r['环比增长']
            style = fmt_red if val > 0 else (fmt_green if val < 0 else fmt_pct)
            ws.write(row, 4, val, style)
            ws.write(row, 5, r[cols[5]], fmt_data)
        chart = workbook.add_chart({'type': 'column'})
        chart.add_series(
            {'name': cols[1], 'categories': [ws.name, 2, 0, len(df) + 1, 0], 'values': [ws.name, 2, 1, len(df) + 1, 1],
             'fill': {'color': chart_color}, 'data_labels': {'value': True, 'position': 'outside_end'}})
        chart.set_size({'width': 600, 'height': 350})
        ws.insert_chart('H3', chart)
        ws.set_column('A:F', 15)

    if not df_in.empty:
        ws1 = workbook.add_worksheet('入库分析')
        write_sheet(ws1, '周度入库统计', df_in, fmt_h_blue, '#4472C4')

    if not df_out.empty:
        ws2 = workbook.add_worksheet('出库分析')
        write_sheet(ws2, '周度出库统计', df_out, fmt_h_org, '#ED7D31')

    if not df_stock.empty:
        ws3 = workbook.add_worksheet('在库分析')
        ws3.merge_range('A1:C1', '当前实时在库库存', fmt_h_grn)
        ws3.write_row('A2', ['客户', '当前在库(片)', '当前在库(批)'], fmt_h_grn)
        df_sorted = df_stock.sort_values(by='当前在库(片)', ascending=False)
        for i, r in enumerate(df_sorted.to_dict('records')):
            ws3.write_row(i + 2, 0, [r['客户'], r['当前在库(片)'], r['当前在库(批)']], fmt_data)
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
            'name': '库存分布',
            'categories': ['在库分析', 2, 0, len(df_stock) + 1, 0],
            'values': ['在库分析', 2, 1, len(df_stock) + 1, 1],
            'data_labels': {'value': False, 'percentage': True, 'position': 'outside_end', 'leader_lines': True,
                            'num_format': '0.00%'}
        })
        ws3.insert_chart('E3', chart)
        ws3.set_column('A:C', 18)

    if df_month is not None and not df_month.empty:
        ws4 = workbook.add_worksheet('月度趋势')
        cols = ['月份_Str', '入库总数', '出库总数']
        ws4.write_row('A1', cols, fmt_h_blue)
        for i, r in enumerate(df_month.to_dict('records')):
            ws4.write_row(i + 1, 0, [r['月份_Str'], r['入库总数'], r['出库总数']], fmt_data)
        ws4.set_column('A:C', 15)
        chart_line = workbook.add_chart({'type': 'line'})
        l_r = len(df_month) + 1
        chart_line.add_series(
            {'name': '入库总数', 'categories': ['月度趋势', 1, 0, l_r, 0], 'values': ['月度趋势', 1, 1, l_r, 1],
             'line': {'color': '#4472C4', 'width': 2.25}, 'marker': {'type': 'circle', 'size': 6}})
        chart_line.add_series(
            {'name': '出库总数', 'categories': ['月度趋势', 1, 0, l_r, 0], 'values': ['月度趋势', 1, 2, l_r, 2],
             'line': {'color': '#ED7D31', 'width': 2.25}, 'marker': {'type': 'square', 'size': 6}})
        chart_line.set_title({'name': '全厂月度出入库总趋势'})
        chart_line.set_size({'width': 800, 'height': 400})
        ws4.insert_chart('E2', chart_line)

    workbook.close()
    return output.getvalue()


def download_search_report(search_df, sheet_headers, t_s=None, t_e=None):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    font = '微软雅黑'
    fmt_header = workbook.add_format(
        {'bold': True, 'bg_color': '#2C3E50', 'font_color': 'white', 'border': 1, 'align': 'center',
         'valign': 'vcenter', 'font_name': font})
    fmt_cell = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_name': font})
    fmt_summary_k = workbook.add_format({'bold': True, 'bg_color': '#D5F5E3', 'border': 1, 'font_name': font})
    fmt_summary_v = workbook.add_format({'border': 1, 'font_name': font, 'align': 'left'})

    total_count = len(search_df)
    in_range_count = 0;
    out_range_count = 0;
    stock_count = 0

    if t_s and t_e:
        ts = pd.to_datetime(t_s);
        te = pd.to_datetime(t_e) + timedelta(days=0.99999)
        in_range_count = len(search_df[(search_df['_sys_in_date'] >= ts) & (search_df['_sys_in_date'] <= te)])
        out_range_count = len(search_df[(search_df['_sys_out_date'] >= ts) & (search_df['_sys_out_date'] <= te)])
    stock_count = len(search_df[pd.isna(search_df['_sys_out_date'])])

    sources = search_df['_sys_source_key'].unique()
    for source_key in sources:
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', str(source_key))[:30]
        ws = workbook.add_worksheet(safe_name)

        ws.write('A1', '统计周期:', fmt_summary_k)
        ws.write('B1', f"{t_s} ~ {t_e}" if t_s else "未指定", fmt_summary_v)
        ws.write('A2', '区间入库数:', fmt_summary_k)
        ws.write('B2', in_range_count, fmt_summary_v)
        ws.write('C2', '区间出库数:', fmt_summary_k)
        ws.write('D2', out_range_count, fmt_summary_v)
        ws.write('A3', '当前在库数:', fmt_summary_k)
        ws.write('B3', stock_count, fmt_summary_v)

        sub_df = search_df[search_df['_sys_source_key'] == source_key]
        original_cols = sheet_headers.get(source_key, [])
        if not original_cols: original_cols = [c for c in sub_df.columns if not c.startswith('_sys_')]

        start_row = 4
        ws.write_row(start_row, 0, original_cols, fmt_header)
        col_widths = {i: len(str(col)) * 2 for i, col in enumerate(original_cols)}
        for i, row in enumerate(sub_df[original_cols].values):
            clean_row = []
            for j, item in enumerate(row):
                val_str = item.strftime('%Y-%m-%d') if isinstance(item, pd.Timestamp) else str(item) if pd.notna(
                    item) else ""
                clean_row.append(val_str)
                col_widths[j] = max(col_widths.get(j, 0), len(val_str))
            ws.write_row(start_row + 1 + i, 0, clean_row, fmt_cell)
        for i, width in col_widths.items(): ws.set_column(i, i, min(width * 1.2 + 2, 50))

    workbook.close()
    return output.getvalue()


def download_stocktake_report(df_stocktake):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    fmt_h = workbook.add_format({'bold': True, 'bg_color': '#FFC000', 'border': 1, 'align': 'center'})
    fmt_d = workbook.add_format({'border': 1, 'align': 'center'})
    ws = workbook.add_worksheet("盘点清单")
    cols = ['客户', '库位', '料盒号', '账面片数', '实盘片数(请填写)', '盘点结果']
    ws.write_row('A1', cols, fmt_h)
    df_sorted = df_stocktake.sort_values(by=['客户', '库位', '料盒号'])
    for i, r in enumerate(df_sorted.to_dict('records')):
        ws.write(i + 1, 0, r['客户'], fmt_d)
        ws.write(i + 1, 1, r['库位'], fmt_d)
        ws.write(i + 1, 2, r['料盒号'], fmt_d)
        ws.write(i + 1, 3, r['账面片数'], fmt_d)
        ws.write(i + 1, 4, '', fmt_d)
        ws.write_formula(i + 1, 5, f'=IF(E{i + 2}="","",IF(E{i + 2}=D{i + 2},"OK","差异"))', fmt_d)
    ws.set_column('A:C', 15)
    ws.set_column('F:F', 10)
    workbook.close()
    return output.getvalue()


def download_stock_report_simple(stock_pivot, chart_data):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    fmt_h = workbook.add_format(
        {'bold': True, 'bg_color': '#34495E', 'font_color': 'white', 'border': 1, 'align': 'center'})
    fmt_d = workbook.add_format({'border': 1, 'align': 'center'})
    ws = workbook.add_worksheet("库存看板")
    flat_stock = stock_pivot.reset_index()
    ws.write_row(0, 0, list(flat_stock.columns), fmt_h)
    for i, row in enumerate(flat_stock.values): ws.write_row(i + 1, 0, row, fmt_d)
    ws.set_column(0, len(flat_stock.columns), 15)
    workbook.close()
    return output.getvalue()


def download_cassette_report(cassette_summary, cassette_details):
    output, workbook, fmts = create_workbook_base()
    ws = workbook.add_worksheet("料盒分析")
    headers = list(cassette_summary.columns)
    ws.write_row(0, 0, headers)
    for i, row in enumerate(cassette_summary.values): ws.write_row(i + 1, 0, row)
    start_row = len(cassette_summary) + 4
    ws.write(start_row, 0, "详细明细")
    export_cols = ['_sys_client', '_sys_cass_type', '_sys_cass_prefix', '_sys_cassette', 'Wafer_Count']
    ws.write_row(start_row + 1, 0, export_cols)
    for i, row in enumerate(cassette_details[export_cols].values): ws.write_row(start_row + 2 + i, 0, row)
    workbook.close()
    return output.getvalue()


def create_workbook_base():
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    formats = {'title': workbook.add_format({'bold': True}), 'header': workbook.add_format({'bold': True}),
               'cell': workbook.add_format({'border': 1})}
    return output, workbook, formats


# ================= 6. 界面逻辑 =================

st.set_page_config(page_title="智能芯片台账 ProMax", layout="wide", page_icon="🐱")
inject_custom_css()
st.title("🐱 智能芯片台账系统 (Ultimate)")

if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
    st.session_state.df_master = pd.DataFrame()
    st.session_state.stock_pivot = pd.DataFrame()
    st.session_state.chart_data = pd.DataFrame()
    st.session_state.cassette_summary = pd.DataFrame()
    st.session_state.cassette_details = pd.DataFrame()
    st.session_state.search_result = pd.DataFrame()
    st.session_state.sheet_headers = {}
    st.session_state.df_stocktake = pd.DataFrame()

with st.sidebar:
    st.header("📂 数据加载")
    password = st.text_input("🔑 全局密码", type="password")

    st.markdown("---")
    st.markdown("#### ⚙️ 高级设置")
    # 🟢 默认关闭灰色识别，响应“不看灰色背景”
    use_grey_logic = st.checkbox("启用灰色背景识别 (针对919等台账)", value=False,
                                 help="默认不看灰色背景。\n如果您的台账（如919）依赖灰色来表示出库，请勾选此项。")
    if use_grey_logic:
        st.warning("⚠️ 已启用灰色背景识别。灰色行将被视为【已出库】。")
    else:
        st.success("✅ 仅根据【领用时间】和【备注关键词】判断出库。")
    st.markdown("---")

    uploaded_files = st.file_uploader("上传 Excel 文件", accept_multiple_files=True, type=['xlsx', 'xls', 'xlsm'])

if uploaded_files:
    # 🟢 传入 use_grey_logic 参数
    df, headers_map = process_uploaded_files(uploaded_files, password, use_grey_logic)
    if not df.empty:
        st.session_state.data_loaded = True
        st.session_state.df_master = df
        st.session_state.sheet_headers = headers_map

        df_stock = df[pd.isna(df['_sys_out_date'])].copy()  # 仅在库
        if not df_stock.empty:
            st.session_state.stock_pivot = df_stock.pivot_table(index='_sys_client', columns='_sys_chip_type',
                                                                values='_sys_wafer', aggfunc='count', fill_value=0)
            st.session_state.chart_data = df_stock.groupby(['_sys_client', '_sys_chip_type']).size().reset_index(
                name='Count')
            # 🟢 盘点清单逻辑
            stocktake = df_stock.groupby(['_sys_client', '_sys_location', '_sys_cassette']).agg(
                账面片数=('_sys_wafer', 'count')
            ).reset_index()
            stocktake.columns = ['客户', '库位', '料盒号', '账面片数']
            st.session_state.df_stocktake = stocktake

        df_c = df[(~df['_sys_cassette'].isin(['-', 'nan', '', '未知/空'])) & (pd.isna(df['_sys_out_date']))]
        if not df_c.empty:
            st.session_state.cassette_details = df_c.groupby(
                ['_sys_client', '_sys_cass_type', '_sys_cass_prefix', '_sys_cassette']).agg(
                Wafer_Count=('_sys_wafer', 'count')).reset_index()
            st.session_state.cassette_summary = st.session_state.cassette_details.groupby(
                ['_sys_cass_type', '_sys_cass_prefix']).agg(料盒总数=('_sys_cassette', 'nunique'),
                                                            涉及客户=('_sys_client', 'unique')).reset_index()
            st.session_state.cassette_summary['涉及客户'] = st.session_state.cassette_summary['涉及客户'].apply(
                lambda x: ', '.join(x))
        st.success(f"✅ 数据加载完成！共 {len(df)} 条记录")

# Tab 页面
tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 库存看板", "🔎 组合查询", "📦 料盒分析", "📋 现场盘点", "📈 统计分析"])

with tab1:
    if st.session_state.data_loaded:
        c1, c2 = st.columns([1, 1.5])
        with c1:
            if not st.session_state.stock_pivot.empty:
                st.dataframe(st.session_state.stock_pivot, width="stretch")
                st.download_button("📥 下载库存报表", download_stock_report_simple(st.session_state.stock_pivot,
                                                                                  st.session_state.chart_data),
                                   "库存.xlsx")
        with c2:
            if not st.session_state.chart_data.empty:
                # 🟢 优化颜色
                chart = alt.Chart(st.session_state.chart_data).mark_bar().encode(
                    x=alt.X('_sys_client', title='客户'),
                    y=alt.Y('Count', title='在库数量'),
                    color=alt.Color('_sys_chip_type', scale=alt.Scale(domain=['Real', 'Dummy', '不可回货'],
                                                                      range=['#1F77B4', '#7F7F7F', '#D62728']))
                ).interactive()
                st.altair_chart(chart, use_container_width=True)
    else:
        st.markdown('<div class="info-box">👋 请先在左侧上传 Excel 数据。</div>', unsafe_allow_html=True)

with tab2:
    if st.session_state.data_loaded:
        with st.form("qry"):
            c1, c2, c3, c4 = st.columns(4)
            clients = sorted(
                st.session_state.df_master['_sys_client'].unique()) if not st.session_state.df_master.empty else []
            sel_client = c1.multiselect("1. 客户", clients)
            sel_status = c2.selectbox("2. 状态", ["全部", "在库", "已出库"])
            q_diff = c3.text_input("3. 扩散批 (模糊)")
            q_wafer = c4.text_input("4. 芯片号 (模糊)")
            d1, d2, d3 = st.columns(3)
            q_cass = d1.text_input("5. 料盒号")
            q_loc = d2.text_input("6. 库位号")
            q_date = d3.date_input("7. 时间范围", [])

            if st.form_submit_button("🚀 执行查询"):
                res = st.session_state.df_master.copy()
                if sel_client: res = res[res['_sys_client'].isin(sel_client)]
                if sel_status == "在库":
                    res = res[pd.isna(res['_sys_out_date'])]
                elif sel_status == "已出库":
                    res = res[pd.notna(res['_sys_out_date'])]

                if q_diff: res = res[res['_sys_diff_lot'].str.contains(q_diff, na=False, case=False)]
                if q_wafer: res = res[res['_sys_wafer'].str.contains(q_wafer, na=False, case=False)]
                if q_cass: res = res[res['_sys_cassette'].str.contains(q_cass, na=False, case=False)]
                if q_loc: res = res[res['_sys_location'].str.contains(q_loc, na=False, case=False)]

                if len(q_date) == 2:
                    s, e = pd.to_datetime(q_date[0]), pd.to_datetime(q_date[1]) + timedelta(days=0.99999)
                    res = res[((res['_sys_in_date'] >= s) & (res['_sys_in_date'] <= e)) |
                              ((res['_sys_out_date'] >= s) & (res['_sys_out_date'] <= e))]

                st.session_state.search_result = res
                st.session_state.query_dates = (q_date[0], q_date[1]) if len(q_date) == 2 else (None, None)

        if not st.session_state.search_result.empty:
            st.dataframe(st.session_state.search_result)
            q_start, q_end = st.session_state.get('query_dates', (None, None))
            st.download_button("📥 下载查询结果 (含统计)",
                               download_search_report(st.session_state.search_result, st.session_state.sheet_headers,
                                                      q_start, q_end),
                               f"查询结果_{datetime.now().strftime('%H%M')}.xlsx")
    else:
        st.markdown('<div class="info-box">👋 暂无数据。</div>', unsafe_allow_html=True)

with tab3:
    if st.session_state.data_loaded:
        c1, c2 = st.columns([1, 2])
        with c1:
            st.dataframe(st.session_state.cassette_summary, width="stretch")
            opts = ["(全部)"] + sorted(list(set([f"{r['_sys_cass_type']} ({r['_sys_cass_prefix']})" for _, r in
                                                 st.session_state.cassette_summary.iterrows()])))
            sel_comb = st.selectbox("筛选类型:", opts)
        with c2:
            show_df = st.session_state.cassette_details
            if sel_comb != "(全部)":
                t = sel_comb.split(' (')[0];
                p = sel_comb.split(' (')[1][:-1]
                show_df = show_df[(show_df['_sys_cass_type'] == t) & (show_df['_sys_cass_prefix'] == p)]
            st.dataframe(show_df, width="stretch")
            st.download_button("下载料盒", download_cassette_report(st.session_state.cassette_summary, show_df),
                               "料盒.xlsx")
    else:
        st.markdown('<div class="info-box">👋 暂无数据。</div>', unsafe_allow_html=True)

with tab4:
    if st.session_state.data_loaded:
        st.subheader("📋 现场实物盘点清单")
        if not st.session_state.df_stocktake.empty:
            c1, c2 = st.columns(2)
            all_clients = ["(全部)"] + sorted(st.session_state.df_stocktake['客户'].unique().tolist())
            sel_client_stock = c1.selectbox("👤 筛选客户:", all_clients)
            loc_filter = c2.text_input("🔍 搜索库位 (例如 E08):").strip().upper()
            display_df = st.session_state.df_stocktake
            if sel_client_stock != "(全部)": display_df = display_df[display_df['客户'] == sel_client_stock]
            if loc_filter: display_df = display_df[display_df['库位'].str.contains(loc_filter, na=False)]
            st.dataframe(display_df, width="stretch", height=500)
            st.download_button("📥 下载盘点单", download_stocktake_report(display_df),
                               f"盘点单_{datetime.now().strftime('%Y%m%d')}.xlsx")
        else:
            st.warning("无在库数据")
    else:
        st.markdown('<div class="info-box">👋 暂无数据。</div>', unsafe_allow_html=True)

with tab5:
    if st.session_state.data_loaded:
        st.subheader("📈 统计分析")
        c1, c2, c3 = st.columns([1, 1, 1])
        start_d = c1.date_input("开始", value=datetime.today() - timedelta(days=7))
        end_d = c2.date_input("结束", value=datetime.today())
        if c3.button("计算"):
            df_in, df_out, df_stk_stat, df_month = calculate_statistics(st.session_state.df_master, start_d, end_d)
            report = create_statistical_excel(df_in, df_out, df_stk_stat, df_month, start_d, end_d)
            st.download_button("下载统计报表", report, "统计.xlsx")
    else:
        st.markdown('<div class="info-box">👋 暂无数据。</div>', unsafe_allow_html=True)